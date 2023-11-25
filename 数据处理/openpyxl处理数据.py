from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd


# 简单测试
def demo():
    wb = Workbook()

    # grab the active worksheet
    ws = wb.active

    # Data can be assigned directly to cells
    ws['A1'] = 42

    # Rows can also be appended
    ws.append([1, 2, 3])

    # Python types will automatically be converted
    import datetime

    ws['A2'] = datetime.datetime.now()

    # Save the file
    wb.save("sample.xlsx")


def init():
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active


def add():
    # 实例化
    wb = Workbook()
    # 激活 worksheet
    ws = wb.active
    # 方式一：数据可以直接分配到单元格中(可以输入公式)
    ws['A1'] = 42
    # 方式二：可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
    ws.append([1, 2, 3])
    # 方式三：Python 类型会被自动转换
    # ws['A3'] = datetime.datetime.now().strftime("%Y-%m-%d")
    wb.save("好好好.xlsx")


def get_chart():
    wb = load_workbook('职工表.xlsx')
    # 激活 worksheet
    ws = wb.active
    # 显示所有表名
    print(wb.sheetnames)
    # 遍历所有表
    for sheet in wb:
        print(sheet.title)

    # 方法一
    c = ws['A4']
    print("c = {}".format(c))
    # 方法二：row 行；column 列
    d = ws.cell(row=4, column=2, value=10)
    print("d = {}".format(d))
    # # 方法三：只要访问就创建
    # for i in range(1, 101):
    #     ...
    #     for j in range(1, 101):
    #         ...
    #     ws.cell(row=i, column=j)


def openpyxl_demo():
    from openpyxl import load_workbook

    # 加载工作簿
    wb = load_workbook(filename='职工表.xlsx')

    # 获取工作表
    ws = wb.active

    # 获取行和列的数量
    max_row = ws.max_row
    max_col = ws.max_column

    # 遍历工作表中的每一个单元格，并打印其值
    for row in range(1, max_row):
        for col in range(1, max_col):
            cell = ws.cell(row=row, column=col)
            print(cell.value, end=' ')
        print('\n')


def openpyxl_demo():
    # 加载工作簿
    wb = load_workbook(filename='职工表.xlsx')

    # 获取工作表
    ws = wb.active

    # 获取行和列的数量
    max_row = ws.max_row
    max_col = ws.max_column

    # 遍历工作表中的每一个单元格，并打印其值
    for row in range(1, max_row):
        for col in range(1, max_col):
            cell = ws.cell(row=row, column=col)
            print(cell.value, end=' ')
        print('\n')

    ws.append([8, "邓几把", 24, "流氓"])
    ws.append([9, "邓几把", 24, "流氓"])
    ws.append([10, "邓几把", 24, "流氓"])

    wb.save('职工表.xlsx')


def pandas_demo():
    # 读取Excel文件
    df = pd.read_excel('职工表.xlsx')

    # 打印数据框的内容
    print(df)


if __name__ == '__main__':
    openpyxl_demo()
    pandas_demo()
